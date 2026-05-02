"""Test and demo for ExcelTemplateBuilder

This script tests all four template types with sample data and generates
demo Excel files in the Downloads folder.

Run with: python -m analytics.exporters.test_excel_templates
"""

import sys
import os
from pathlib import Path
from datetime import datetime, timedelta

# Set UTF-8 encoding for Windows console
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from analytics.exporters.excel_exporter import ExcelExporter
from analytics.exporters.excel_templates import ExcelTemplateBuilder


def create_sample_report_data(days: int = 30) -> dict:
    """Create realistic sample report data for testing

    Args:
        days: Number of days for the report period

    Returns:
        dict: Sample report data matching ReportGenerator structure
    """
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)

    return {
        "metadata": {
            "generated_at": datetime.now().isoformat(),
            "report_type": "detailed",
            "date_range": {
                "start": start_date.strftime("%Y-%m-%d"),
                "end": end_date.strftime("%Y-%m-%d"),
                "days": days
            }
        },
        "sections": [
            {
                "title": "Overview",
                "metrics": {
                    "total_sessions": 145,
                    "total_skill_invocations": 423,
                    "total_tokens": 1250000,
                    "total_cost_usd": 15.75
                }
            },
            {
                "title": "Top Skills",
                "metrics": {
                    "top_skills": [
                        {"skill_name": "dream-studio:core", "invocations": 89, "success_rate": 0.95},
                        {"skill_name": "dream-studio:quality", "invocations": 67, "success_rate": 0.92},
                        {"skill_name": "dream-studio:security", "invocations": 45, "success_rate": 0.88},
                        {"skill_name": "dream-studio:domains", "invocations": 38, "success_rate": 0.90},
                        {"skill_name": "dream-studio:analyze", "invocations": 32, "success_rate": 0.94},
                        {"skill_name": "dream-studio:career", "invocations": 28, "success_rate": 0.91},
                        {"skill_name": "simplify", "invocations": 24, "success_rate": 0.89},
                        {"skill_name": "graphify", "invocations": 20, "success_rate": 0.93},
                        {"skill_name": "huashu-design", "invocations": 18, "success_rate": 0.87},
                        {"skill_name": "claude-api", "invocations": 15, "success_rate": 0.96},
                    ],
                    "success_rate_overall": 0.93
                }
            },
            {
                "title": "Token Usage",
                "metrics": {
                    "by_model": {
                        "claude-sonnet-4-5": 850000,
                        "claude-haiku-3-5": 350000,
                        "claude-opus-4": 50000
                    },
                    "daily_average": 41667,
                    "by_skill": {
                        "dream-studio:core": 450000,
                        "dream-studio:quality": 320000,
                        "dream-studio:security": 180000,
                        "other": 300000
                    },
                    "by_project": {
                        "dream-studio": 650000,
                        "dreamysuite": 400000,
                        "analytics": 200000
                    }
                }
            },
            {
                "title": "Session Analytics",
                "metrics": {
                    "by_project": {
                        "dream-studio": 78,
                        "dreamysuite": 45,
                        "analytics": 22
                    },
                    "day_of_week": {
                        "Monday": 25,
                        "Tuesday": 30,
                        "Wednesday": 22,
                        "Thursday": 28,
                        "Friday": 20,
                        "Saturday": 10,
                        "Sunday": 10
                    },
                    "avg_duration_minutes": 45.5,
                    "outcomes": {
                        "success": 120,
                        "partial": 20,
                        "failed": 5
                    }
                }
            },
            {
                "title": "Model Usage",
                "metrics": {
                    "by_model": {
                        "claude-sonnet-4-5": 89,
                        "claude-haiku-3-5": 45,
                        "claude-opus-4": 11
                    },
                    "model_switches": [
                        {"from": "sonnet", "to": "haiku", "count": 23},
                        {"from": "haiku", "to": "sonnet", "count": 18},
                        {"from": "sonnet", "to": "opus", "count": 5}
                    ]
                }
            },
            {
                "title": "Lessons Learned",
                "metrics": {
                    "total_lessons": 12,
                    "by_source": {
                        "build": 5,
                        "debug": 4,
                        "deploy": 2,
                        "review": 1
                    },
                    "by_status": {
                        "active": 10,
                        "archived": 2
                    },
                    "recent_lessons": [
                        {
                            "lesson": "Always check for open PRs before pushing to avoid conflicts",
                            "source": "build",
                            "status": "active"
                        },
                        {
                            "lesson": "Use Test-Path instead of Glob immediately after subagent writes",
                            "source": "debug",
                            "status": "active"
                        },
                        {
                            "lesson": "Wave-based parallel subagent dispatch is highly efficient",
                            "source": "build",
                            "status": "active"
                        },
                        {
                            "lesson": "Trace all call sites before removing guard clauses",
                            "source": "debug",
                            "status": "active"
                        },
                        {
                            "lesson": "Complete and commit each task independently for context resilience",
                            "source": "build",
                            "status": "active"
                        }
                    ]
                }
            },
            {
                "title": "Workflow Analytics",
                "metrics": {
                    "total_runs": 87,
                    "by_workflow": {
                        "build-lifecycle": 45,
                        "security-scan": 23,
                        "quality-audit": 12,
                        "deploy-pipeline": 7
                    },
                    "by_status": {
                        "completed": 75,
                        "failed": 8,
                        "in_progress": 4
                    },
                    "success_rate": 0.86,
                    "avg_completion_time": 12.5
                }
            },
            {
                "title": "Detailed Skill Metrics",
                "metrics": {
                    "by_skill": {
                        "dream-studio:core": {
                            "invocations": 89,
                            "success_rate": 0.95,
                            "avg_duration": 8.5
                        },
                        "dream-studio:quality": {
                            "invocations": 67,
                            "success_rate": 0.92,
                            "avg_duration": 6.2
                        }
                    },
                    "failures": [
                        {"skill": "dream-studio:security", "reason": "timeout", "timestamp": "2026-04-28"},
                        {"skill": "simplify", "reason": "parse_error", "timestamp": "2026-04-27"}
                    ]
                }
            }
        ]
    }


def create_historical_report_data(days: int = 30) -> dict:
    """Create historical report data for comparison testing

    Args:
        days: Number of days for the report period

    Returns:
        dict: Historical report data with lower values
    """
    end_date = datetime.now() - timedelta(days=days)
    start_date = end_date - timedelta(days=days)

    return {
        "metadata": {
            "generated_at": end_date.isoformat(),
            "report_type": "detailed",
            "date_range": {
                "start": start_date.strftime("%Y-%m-%d"),
                "end": end_date.strftime("%Y-%m-%d"),
                "days": days
            }
        },
        "sections": [
            {
                "title": "Overview",
                "metrics": {
                    "total_sessions": 120,
                    "total_skill_invocations": 350,
                    "total_tokens": 1000000,
                    "total_cost_usd": 12.50
                }
            },
            {
                "title": "Top Skills",
                "metrics": {
                    "success_rate_overall": 0.90
                }
            },
            {
                "title": "Token Usage",
                "metrics": {
                    "daily_average": 33333
                }
            },
            {
                "title": "Session Analytics",
                "metrics": {
                    "avg_duration_minutes": 50.0
                }
            },
            {
                "title": "Workflow Analytics",
                "metrics": {
                    "success_rate": 0.82
                }
            }
        ]
    }


def test_summary_dashboard():
    """Test Summary Dashboard template"""
    print("\n=== Testing Summary Dashboard Template ===")

    try:
        exporter = ExcelExporter()
        builder = ExcelTemplateBuilder()

        # Create workbook
        wb = exporter.create_workbook()

        # Generate report data
        report = create_sample_report_data(30)

        # Build template
        builder.build_summary_dashboard(report, wb)

        # Save
        output_path = "C:/Users/Dannis Seay/Downloads/analytics_summary_dashboard.xlsx"
        result = exporter.save_workbook(wb, output_path)

        if result["success"]:
            print(f"✓ Summary Dashboard created successfully")
            print(f"  Path: {result['path']}")
        else:
            print(f"✗ Failed: {result['error']}")

    except Exception as e:
        print(f"✗ Error: {str(e)}")


def test_raw_data_dump():
    """Test Raw Data Dump template"""
    print("\n=== Testing Raw Data Dump Template ===")

    try:
        exporter = ExcelExporter()
        builder = ExcelTemplateBuilder()

        # Create workbook
        wb = exporter.create_workbook()

        # Generate report data
        report = create_sample_report_data(30)

        # Build template
        builder.build_raw_data_dump(report, wb)

        # Save
        output_path = "C:/Users/Dannis Seay/Downloads/analytics_raw_data_dump.xlsx"
        result = exporter.save_workbook(wb, output_path)

        if result["success"]:
            print(f"✓ Raw Data Dump created successfully")
            print(f"  Path: {result['path']}")
            print(f"  Sheets: {len(wb.sheetnames)}")
        else:
            print(f"✗ Failed: {result['error']}")

    except Exception as e:
        print(f"✗ Error: {str(e)}")


def test_trend_analysis():
    """Test Trend Analysis template"""
    print("\n=== Testing Trend Analysis Template ===")

    try:
        exporter = ExcelExporter()
        builder = ExcelTemplateBuilder()

        # Create workbook
        wb = exporter.create_workbook()

        # Generate report data
        report = create_sample_report_data(30)

        # Build template
        builder.build_trend_analysis(report, wb)

        # Save
        output_path = "C:/Users/Dannis Seay/Downloads/analytics_trend_analysis.xlsx"
        result = exporter.save_workbook(wb, output_path)

        if result["success"]:
            print(f"✓ Trend Analysis created successfully")
            print(f"  Path: {result['path']}")
        else:
            print(f"✗ Failed: {result['error']}")

    except Exception as e:
        print(f"✗ Error: {str(e)}")


def test_comparison():
    """Test Comparison template"""
    print("\n=== Testing Comparison Template ===")

    try:
        exporter = ExcelExporter()
        builder = ExcelTemplateBuilder()

        # Create workbook
        wb = exporter.create_workbook()

        # Generate current and historical data
        current_report = create_sample_report_data(30)
        historical_report = create_historical_report_data(30)

        # Build template
        builder.build_comparison(current_report, wb, historical_report)

        # Save
        output_path = "C:/Users/Dannis Seay/Downloads/analytics_comparison.xlsx"
        result = exporter.save_workbook(wb, output_path)

        if result["success"]:
            print(f"✓ Comparison created successfully")
            print(f"  Path: {result['path']}")
        else:
            print(f"✗ Failed: {result['error']}")

    except Exception as e:
        print(f"✗ Error: {str(e)}")


def test_all_templates():
    """Test all templates in one comprehensive workbook"""
    print("\n=== Testing All Templates Combined ===")

    try:
        exporter = ExcelExporter()
        builder = ExcelTemplateBuilder()

        # Create workbook
        wb = exporter.create_workbook()

        # Generate data
        current_report = create_sample_report_data(30)
        historical_report = create_historical_report_data(30)

        # Build all templates
        builder.build_summary_dashboard(current_report, wb)
        builder.build_raw_data_dump(current_report, wb)
        builder.build_trend_analysis(current_report, wb)
        builder.build_comparison(current_report, wb, historical_report)

        # Save
        output_path = "C:/Users/Dannis Seay/Downloads/analytics_complete_report.xlsx"
        result = exporter.save_workbook(wb, output_path)

        if result["success"]:
            print(f"✓ Complete Report created successfully")
            print(f"  Path: {result['path']}")
            print(f"  Total Sheets: {len(wb.sheetnames)}")
            print(f"  Sheets: {', '.join(wb.sheetnames)}")
        else:
            print(f"✗ Failed: {result['error']}")

    except Exception as e:
        print(f"✗ Error: {str(e)}")


def main():
    """Run all template tests"""
    print("=" * 60)
    print("Excel Template Builder - Test Suite")
    print("=" * 60)

    # Check for openpyxl
    try:
        import openpyxl
        print("✓ openpyxl is installed")
    except ImportError:
        print("✗ openpyxl is NOT installed")
        print("\nInstall with: pip install openpyxl")
        return

    # Run individual tests
    test_summary_dashboard()
    test_raw_data_dump()
    test_trend_analysis()
    test_comparison()

    # Run combined test
    test_all_templates()

    print("\n" + "=" * 60)
    print("All tests completed!")
    print("Check C:/Users/Dannis Seay/Downloads/ for generated files")
    print("=" * 60)


if __name__ == "__main__":
    main()

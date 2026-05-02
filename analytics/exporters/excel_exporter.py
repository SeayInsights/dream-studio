"""Excel exporter for analytics reports

Exports analytics reports to Excel format with multiple sheets, charts, and formatting.
Supports openpyxl (full features), xlsxwriter (write-only fallback), and CSV (final fallback).

Example usage:
    from analytics.exporters.excel_exporter import ExcelExporter

    # Create exporter
    exporter = ExcelExporter()

    # Export report data
    result = exporter.export_to_excel(
        data=report_data,
        output_path="C:/Users/Dannis Seay/Downloads/analytics_report.xlsx"
    )

    if result["success"]:
        print(f"Report exported to: {result['path']}")
    else:
        print(f"Export failed: {result['error']}")
"""

import sys
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import csv

# Try to import Excel libraries in order of preference
HAS_OPENPYXL = False
HAS_XLSXWRITER = False

try:
    import openpyxl
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    pass

try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    pass


class ExcelExporter:
    """Export analytics reports to Excel format with charts and formatting"""

    def __init__(self):
        """Initialize exporter with library availability check"""
        self.library = self._detect_library()

    def _detect_library(self) -> str:
        """Detect which Excel library is available

        Returns:
            str: "openpyxl", "xlsxwriter", or "csv"
        """
        if HAS_OPENPYXL:
            return "openpyxl"
        elif HAS_XLSXWRITER:
            return "xlsxwriter"
        else:
            return "csv"

    def export_to_excel(
        self,
        data: Dict[str, Any],
        output_path: str
    ) -> Dict[str, Any]:
        """Export report data to Excel file

        Args:
            data: Report data dictionary with structure:
                {
                    "id": str,
                    "name": str,
                    "type": str,
                    "metadata": dict,
                    "metrics": dict,
                    "insights": dict,
                    "recommendations": list,
                    "charts": list (optional),
                    "generated_at": datetime
                }
            output_path: Full path to output Excel file

        Returns:
            dict: Result with keys:
                - success: bool
                - path: str (if success)
                - library: str (which library was used)
                - error: str (if failure)
        """
        try:
            # Validate data structure
            validation_error = self._validate_data(data)
            if validation_error:
                return {
                    "success": False,
                    "error": f"Data validation failed: {validation_error}"
                }

            # Ensure output directory exists
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)

            # Export based on available library
            if self.library == "openpyxl":
                result_path = self._export_openpyxl(data, output_file)
            elif self.library == "xlsxwriter":
                result_path = self._export_xlsxwriter(data, output_file)
            else:
                result_path = self._export_csv(data, output_file)

            return {
                "success": True,
                "path": str(result_path),
                "library": self.library
            }

        except PermissionError:
            return {
                "success": False,
                "error": f"Permission denied writing to {output_path}. File may be open in Excel."
            }
        except Exception as e:
            return {
                "success": False,
                "error": f"Export failed: {str(e)}"
            }

    def _validate_data(self, data: Dict[str, Any]) -> Optional[str]:
        """Validate report data structure

        Args:
            data: Report data to validate

        Returns:
            str: Error message if validation fails, None if valid
        """
        required_keys = ["id", "name", "type", "metadata", "metrics", "generated_at"]
        for key in required_keys:
            if key not in data:
                return f"Missing required key: {key}"

        if not isinstance(data["metrics"], dict):
            return "metrics must be a dictionary"

        if "insights" in data and not isinstance(data["insights"], dict):
            return "insights must be a dictionary"

        if "recommendations" in data and not isinstance(data["recommendations"], list):
            return "recommendations must be a list"

        return None

    def _export_openpyxl(self, data: Dict[str, Any], output_file: Path) -> Path:
        """Export using openpyxl (full features)

        Args:
            data: Report data
            output_file: Output file path

        Returns:
            Path: Path to created file
        """
        wb = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create sheets
        self._create_summary_sheet_openpyxl(wb, data)
        self._create_metrics_sheet_openpyxl(wb, data)

        if data.get("recommendations"):
            self._create_recommendations_sheet_openpyxl(wb, data)

        if data.get("insights"):
            self._create_insights_sheet_openpyxl(wb, data)

        # Save workbook
        wb.save(output_file)
        return output_file

    def _create_summary_sheet_openpyxl(self, wb: Any, data: Dict[str, Any]) -> None:
        """Create summary sheet with openpyxl

        Args:
            wb: Workbook object
            data: Report data
        """
        ws = wb.create_sheet("Summary", 0)

        # Define styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="left", vertical="center")

        # Title
        ws["A1"] = data["name"]
        ws["A1"].font = header_font
        ws.merge_cells("A1:D1")

        # Metadata
        row = 3
        ws[f"A{row}"] = "Report Details"
        ws[f"A{row}"].font = subheader_font

        row += 1
        metadata_items = [
            ("Report ID", data["id"]),
            ("Type", data["type"]),
            ("Generated", data["generated_at"].strftime("%Y-%m-%d %H:%M:%S")),
            ("Period (days)", data["metadata"].get("days", "N/A")),
        ]

        for label, value in metadata_items:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = str(value)
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Key Metrics Summary
        row += 2
        ws[f"A{row}"] = "Key Metrics"
        ws[f"A{row}"].font = subheader_font
        row += 1

        # Header row
        headers = ["Metric", "Value"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = header_alignment

        row += 1

        # Metrics data
        metrics = data.get("metrics", {})
        for metric_name, metric_value in metrics.items():
            ws[f"A{row}"] = metric_name
            ws[f"B{row}"] = self._format_metric_value(metric_value)
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        # Freeze top row
        ws.freeze_panes = "A4"

    def _create_metrics_sheet_openpyxl(self, wb: Any, data: Dict[str, Any]) -> None:
        """Create detailed metrics sheet with openpyxl

        Args:
            wb: Workbook object
            data: Report data
        """
        ws = wb.create_sheet("Metrics")

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        ws["A1"] = "Metric Name"
        ws["B1"] = "Value"
        ws["C1"] = "Type"
        ws["D1"] = "Details"

        for col in ["A1", "B1", "C1", "D1"]:
            ws[col].font = header_font
            ws[col].fill = header_fill
            ws[col].alignment = header_alignment

        # Populate metrics
        row = 2
        metrics = data.get("metrics", {})

        for metric_name, metric_value in metrics.items():
            ws[f"A{row}"] = metric_name
            ws[f"B{row}"] = self._format_metric_value(metric_value)
            ws[f"C{row}"] = type(metric_value).__name__
            ws[f"D{row}"] = str(metric_value) if isinstance(metric_value, (dict, list)) else ""
            row += 1

        # Auto-adjust columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 40

        # Freeze header
        ws.freeze_panes = "A2"

    def _create_recommendations_sheet_openpyxl(self, wb: Any, data: Dict[str, Any]) -> None:
        """Create recommendations sheet with openpyxl

        Args:
            wb: Workbook object
            data: Report data
        """
        ws = wb.create_sheet("Recommendations")

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Headers
        headers = ["#", "Recommendation", "Priority", "Category"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill

        # Populate recommendations
        row = 2
        recommendations = data.get("recommendations", [])

        for idx, rec in enumerate(recommendations, start=1):
            ws[f"A{row}"] = idx
            ws[f"B{row}"] = rec.get("text", str(rec))
            ws[f"C{row}"] = rec.get("priority", "N/A")
            ws[f"D{row}"] = rec.get("category", "N/A")
            row += 1

        # Auto-adjust columns
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20

        # Freeze header
        ws.freeze_panes = "A2"

    def _create_insights_sheet_openpyxl(self, wb: Any, data: Dict[str, Any]) -> None:
        """Create insights sheet with openpyxl

        Args:
            wb: Workbook object
            data: Report data
        """
        ws = wb.create_sheet("Insights")

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Headers
        ws["A1"] = "Insight Type"
        ws["B1"] = "Description"
        ws["C1"] = "Impact"

        for col in ["A1", "B1", "C1"]:
            ws[col].font = header_font
            ws[col].fill = header_fill

        # Populate insights
        row = 2
        insights = data.get("insights", {})

        for insight_type, insight_data in insights.items():
            ws[f"A{row}"] = insight_type

            if isinstance(insight_data, dict):
                ws[f"B{row}"] = insight_data.get("description", str(insight_data))
                ws[f"C{row}"] = insight_data.get("impact", "N/A")
            else:
                ws[f"B{row}"] = str(insight_data)
                ws[f"C{row}"] = "N/A"

            row += 1

        # Auto-adjust columns
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 15

        # Freeze header
        ws.freeze_panes = "A2"

    def _export_xlsxwriter(self, data: Dict[str, Any], output_file: Path) -> Path:
        """Export using xlsxwriter (write-only, faster)

        Args:
            data: Report data
            output_file: Output file path

        Returns:
            Path: Path to created file
        """
        workbook = xlsxwriter.Workbook(str(output_file))

        # Define formats
        header_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '#366092',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })

        bold_format = workbook.add_format({'bold': True})
        title_format = workbook.add_format({'bold': True, 'font_size': 14})

        # Create sheets
        self._create_summary_sheet_xlsxwriter(workbook, data, header_format, bold_format, title_format)
        self._create_metrics_sheet_xlsxwriter(workbook, data, header_format)

        if data.get("recommendations"):
            self._create_recommendations_sheet_xlsxwriter(workbook, data, header_format)

        if data.get("insights"):
            self._create_insights_sheet_xlsxwriter(workbook, data, header_format)

        workbook.close()
        return output_file

    def _create_summary_sheet_xlsxwriter(
        self,
        workbook: Any,
        data: Dict[str, Any],
        header_format: Any,
        bold_format: Any,
        title_format: Any
    ) -> None:
        """Create summary sheet with xlsxwriter

        Args:
            workbook: Workbook object
            data: Report data
            header_format: Header cell format
            bold_format: Bold cell format
            title_format: Title cell format
        """
        worksheet = workbook.add_worksheet("Summary")

        # Title
        worksheet.write(0, 0, data["name"], title_format)
        worksheet.merge_range("A1:D1", data["name"], title_format)

        # Metadata
        row = 2
        worksheet.write(row, 0, "Report Details", bold_format)

        row += 1
        metadata_items = [
            ("Report ID", data["id"]),
            ("Type", data["type"]),
            ("Generated", data["generated_at"].strftime("%Y-%m-%d %H:%M:%S")),
            ("Period (days)", data["metadata"].get("days", "N/A")),
        ]

        for label, value in metadata_items:
            worksheet.write(row, 0, label, bold_format)
            worksheet.write(row, 1, str(value))
            row += 1

        # Key Metrics
        row += 2
        worksheet.write(row, 0, "Key Metrics", bold_format)
        row += 1

        # Headers
        worksheet.write(row, 0, "Metric", header_format)
        worksheet.write(row, 1, "Value", header_format)
        row += 1

        # Metrics data
        metrics = data.get("metrics", {})
        for metric_name, metric_value in metrics.items():
            worksheet.write(row, 0, metric_name)
            worksheet.write(row, 1, self._format_metric_value(metric_value))
            row += 1

        # Set column widths
        worksheet.set_column("A:A", 30)
        worksheet.set_column("B:B", 20)

        # Freeze panes
        worksheet.freeze_panes(3, 0)

    def _create_metrics_sheet_xlsxwriter(
        self,
        workbook: Any,
        data: Dict[str, Any],
        header_format: Any
    ) -> None:
        """Create metrics sheet with xlsxwriter

        Args:
            workbook: Workbook object
            data: Report data
            header_format: Header cell format
        """
        worksheet = workbook.add_worksheet("Metrics")

        # Headers
        headers = ["Metric Name", "Value", "Type", "Details"]
        for col_idx, header in enumerate(headers):
            worksheet.write(0, col_idx, header, header_format)

        # Populate metrics
        row = 1
        metrics = data.get("metrics", {})

        for metric_name, metric_value in metrics.items():
            worksheet.write(row, 0, metric_name)
            worksheet.write(row, 1, self._format_metric_value(metric_value))
            worksheet.write(row, 2, type(metric_value).__name__)
            worksheet.write(row, 3, str(metric_value) if isinstance(metric_value, (dict, list)) else "")
            row += 1

        # Set column widths
        worksheet.set_column("A:A", 30)
        worksheet.set_column("B:B", 20)
        worksheet.set_column("C:C", 15)
        worksheet.set_column("D:D", 40)

        # Freeze header
        worksheet.freeze_panes(1, 0)

    def _create_recommendations_sheet_xlsxwriter(
        self,
        workbook: Any,
        data: Dict[str, Any],
        header_format: Any
    ) -> None:
        """Create recommendations sheet with xlsxwriter

        Args:
            workbook: Workbook object
            data: Report data
            header_format: Header cell format
        """
        worksheet = workbook.add_worksheet("Recommendations")

        # Headers
        headers = ["#", "Recommendation", "Priority", "Category"]
        for col_idx, header in enumerate(headers):
            worksheet.write(0, col_idx, header, header_format)

        # Populate recommendations
        row = 1
        recommendations = data.get("recommendations", [])

        for idx, rec in enumerate(recommendations, start=1):
            worksheet.write(row, 0, idx)
            worksheet.write(row, 1, rec.get("text", str(rec)))
            worksheet.write(row, 2, rec.get("priority", "N/A"))
            worksheet.write(row, 3, rec.get("category", "N/A"))
            row += 1

        # Set column widths
        worksheet.set_column("A:A", 5)
        worksheet.set_column("B:B", 60)
        worksheet.set_column("C:C", 15)
        worksheet.set_column("D:D", 20)

        # Freeze header
        worksheet.freeze_panes(1, 0)

    def _create_insights_sheet_xlsxwriter(
        self,
        workbook: Any,
        data: Dict[str, Any],
        header_format: Any
    ) -> None:
        """Create insights sheet with xlsxwriter

        Args:
            workbook: Workbook object
            data: Report data
            header_format: Header cell format
        """
        worksheet = workbook.add_worksheet("Insights")

        # Headers
        headers = ["Insight Type", "Description", "Impact"]
        for col_idx, header in enumerate(headers):
            worksheet.write(0, col_idx, header, header_format)

        # Populate insights
        row = 1
        insights = data.get("insights", {})

        for insight_type, insight_data in insights.items():
            worksheet.write(row, 0, insight_type)

            if isinstance(insight_data, dict):
                worksheet.write(row, 1, insight_data.get("description", str(insight_data)))
                worksheet.write(row, 2, insight_data.get("impact", "N/A"))
            else:
                worksheet.write(row, 1, str(insight_data))
                worksheet.write(row, 2, "N/A")

            row += 1

        # Set column widths
        worksheet.set_column("A:A", 25)
        worksheet.set_column("B:B", 60)
        worksheet.set_column("C:C", 15)

        # Freeze header
        worksheet.freeze_panes(1, 0)

    def _export_csv(self, data: Dict[str, Any], output_file: Path) -> Path:
        """Export as CSV (final fallback)

        Args:
            data: Report data
            output_file: Output file path (will change extension to .csv)

        Returns:
            Path: Path to created CSV file
        """
        # Change extension to .csv
        csv_file = output_file.with_suffix(".csv")

        with open(csv_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Title
            writer.writerow([data["name"]])
            writer.writerow([])

            # Metadata
            writer.writerow(["Report Details"])
            writer.writerow(["Report ID", data["id"]])
            writer.writerow(["Type", data["type"]])
            writer.writerow(["Generated", data["generated_at"].strftime("%Y-%m-%d %H:%M:%S")])
            writer.writerow(["Period (days)", data["metadata"].get("days", "N/A")])
            writer.writerow([])

            # Metrics
            writer.writerow(["Metrics"])
            writer.writerow(["Metric", "Value"])

            metrics = data.get("metrics", {})
            for metric_name, metric_value in metrics.items():
                writer.writerow([metric_name, self._format_metric_value(metric_value)])

            writer.writerow([])

            # Recommendations
            if data.get("recommendations"):
                writer.writerow(["Recommendations"])
                writer.writerow(["#", "Recommendation", "Priority", "Category"])

                for idx, rec in enumerate(data["recommendations"], start=1):
                    writer.writerow([
                        idx,
                        rec.get("text", str(rec)),
                        rec.get("priority", "N/A"),
                        rec.get("category", "N/A")
                    ])

                writer.writerow([])

            # Insights
            if data.get("insights"):
                writer.writerow(["Insights"])
                writer.writerow(["Insight Type", "Description", "Impact"])

                insights = data.get("insights", {})
                for insight_type, insight_data in insights.items():
                    if isinstance(insight_data, dict):
                        writer.writerow([
                            insight_type,
                            insight_data.get("description", str(insight_data)),
                            insight_data.get("impact", "N/A")
                        ])
                    else:
                        writer.writerow([insight_type, str(insight_data), "N/A"])

        return csv_file

    def _format_metric_value(self, value: Any) -> str:
        """Format metric value for display

        Args:
            value: Metric value to format

        Returns:
            str: Formatted value
        """
        if isinstance(value, (int, float)):
            if isinstance(value, float):
                return f"{value:,.2f}"
            else:
                return f"{value:,}"
        elif isinstance(value, bool):
            return "Yes" if value else "No"
        elif isinstance(value, (dict, list)):
            return str(value)
        else:
            return str(value)

    def create_workbook(self) -> Any:
        """Create a new workbook object

        Returns:
            Workbook object (openpyxl.Workbook if available, else None)

        Raises:
            RuntimeError: If openpyxl is not installed
        """
        if not HAS_OPENPYXL:
            raise RuntimeError(
                "openpyxl is required for workbook creation. "
                "Install it with: pip install openpyxl"
            )

        wb = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        return wb

    def save_workbook(self, workbook: Any, output_path: str) -> Dict[str, Any]:
        """Save a workbook to file

        Args:
            workbook: openpyxl Workbook object
            output_path: Full path to output Excel file

        Returns:
            dict: Result with keys:
                - success: bool
                - path: str (if success)
                - error: str (if failure)
        """
        try:
            # Ensure output directory exists
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)

            # Save workbook
            workbook.save(output_file)

            return {
                "success": True,
                "path": str(output_file)
            }

        except PermissionError:
            return {
                "success": False,
                "error": f"Permission denied writing to {output_path}. File may be open in Excel."
            }
        except Exception as e:
            return {
                "success": False,
                "error": f"Save failed: {str(e)}"
            }

    def get_installation_instructions(self) -> str:
        """Get installation instructions for Excel libraries

        Returns:
            str: Installation instructions
        """
        if HAS_OPENPYXL:
            return "openpyxl is installed and ready to use."
        elif HAS_XLSXWRITER:
            return "xlsxwriter is installed. For full features (read/write), install openpyxl: pip install openpyxl"
        else:
            return (
                "No Excel libraries installed. Install one of:\n"
                "  - openpyxl (recommended): pip install openpyxl\n"
                "  - xlsxwriter (write-only): pip install xlsxwriter\n"
                "\n"
                "CSV export is available as fallback."
            )


def main():
    """Demo usage of ExcelExporter"""
    from datetime import datetime

    # Sample report data
    sample_data = {
        "id": "report-123",
        "name": "Analytics Report - Q1 2026",
        "type": "executive",
        "metadata": {
            "days": 90,
            "filters": {"department": "engineering"},
            "sections": ["metrics", "insights", "recommendations"]
        },
        "metrics": {
            "total_users": 15234,
            "active_users": 12458,
            "conversion_rate": 0.0823,
            "revenue": 125430.50,
            "churn_rate": 0.0234
        },
        "insights": {
            "user_growth": {
                "description": "User base grew 15% over the quarter",
                "impact": "high"
            },
            "engagement": {
                "description": "Daily active users increased 8%",
                "impact": "medium"
            }
        },
        "recommendations": [
            {
                "text": "Increase marketing spend in Q2 to capitalize on growth",
                "priority": "high",
                "category": "growth"
            },
            {
                "text": "Implement retention campaign for at-risk users",
                "priority": "medium",
                "category": "retention"
            }
        ],
        "generated_at": datetime.now()
    }

    # Create exporter
    exporter = ExcelExporter()

    # Show library info
    print(f"Using library: {exporter.library}")
    print(exporter.get_installation_instructions())
    print()

    # Export report
    output_path = "C:/Users/Dannis Seay/Downloads/sample_analytics_report.xlsx"
    result = exporter.export_to_excel(sample_data, output_path)

    if result["success"]:
        print("Report exported successfully")
        print(f"  Path: {result['path']}")
        print(f"  Library: {result['library']}")
    else:
        print(f"Export failed: {result['error']}")


if __name__ == "__main__":
    main()

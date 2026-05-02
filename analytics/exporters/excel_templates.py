"""Excel template builder for common report types

Provides predefined Excel templates for analytics reports with charts, formatting,
and professional styling using openpyxl.

Template types:
1. Summary Dashboard - KPIs + charts overview
2. Raw Data Dump - All metrics in tabular format
3. Trend Analysis - Time series data with trend charts
4. Comparison - Current vs historical side-by-side

Example usage:
    from analytics.exporters.excel_exporter import ExcelExporter
    from analytics.exporters.excel_templates import ExcelTemplateBuilder
    from analytics.core.reports import ReportGenerator

    # Generate report data
    generator = ReportGenerator()
    report = generator.generate_report(report_type="detailed")

    # Create workbook with template
    exporter = ExcelExporter()
    template_builder = ExcelTemplateBuilder()

    wb = exporter.create_workbook()
    template_builder.build_summary_dashboard(report, wb)
    exporter.save_workbook(wb, "C:/Users/Dannis Seay/Downloads/dashboard.xlsx")
"""

from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import openpyxl
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


class ExcelTemplateBuilder:
    """Build predefined Excel templates for analytics reports"""

    # Color scheme (professional blue palette)
    COLOR_HEADER = "366092"
    COLOR_HEADER_TEXT = "FFFFFF"
    COLOR_KPI_BG = "4472C4"
    COLOR_KPI_TEXT = "FFFFFF"
    COLOR_SUCCESS = "70AD47"
    COLOR_WARNING = "FFC000"
    COLOR_DANGER = "C00000"
    COLOR_NEUTRAL = "7F7F7F"
    COLOR_STRIPE_1 = "FFFFFF"
    COLOR_STRIPE_2 = "F2F2F2"

    def __init__(self):
        """Initialize template builder"""
        pass

    # =============================================================================
    # PUBLIC TEMPLATE METHODS
    # =============================================================================

    def build_summary_dashboard(
        self,
        data: Dict[str, Any],
        workbook: openpyxl.Workbook
    ) -> None:
        """Create summary dashboard template with KPIs and charts

        Args:
            data: Report data from ReportGenerator.generate_report()
            workbook: openpyxl Workbook object to add sheets to
        """
        # Extract metrics from report structure
        metrics = self._extract_metrics_from_report(data)
        metadata = data.get('metadata', {})

        # Sheet 1: KPI Cards
        self._create_kpi_cards_sheet(workbook, metrics, metadata)

        # Sheet 2: Charts
        self._create_charts_sheet(workbook, metrics, metadata)

        # Sheet 3: Top Lists
        self._create_top_lists_sheet(workbook, metrics, metadata)

    def build_raw_data_dump(
        self,
        data: Dict[str, Any],
        workbook: openpyxl.Workbook
    ) -> None:
        """Create raw data dump template with all metrics in tabular format

        Args:
            data: Report data from ReportGenerator.generate_report()
            workbook: openpyxl Workbook object to add sheets to
        """
        metrics = self._extract_metrics_from_report(data)

        # Create a sheet per metric category
        categories = [
            ('Skills', metrics.get('skills', {})),
            ('Tokens', metrics.get('tokens', {})),
            ('Sessions', metrics.get('sessions', {})),
            ('Models', metrics.get('models', {})),
            ('Lessons', metrics.get('lessons', {})),
            ('Workflows', metrics.get('workflows', {}))
        ]

        for category_name, category_data in categories:
            self._create_raw_data_sheet(workbook, category_name, category_data)

    def build_trend_analysis(
        self,
        data: Dict[str, Any],
        workbook: openpyxl.Workbook
    ) -> None:
        """Create trend analysis template with time series data and trend charts

        Args:
            data: Report data from ReportGenerator.generate_report()
            workbook: openpyxl Workbook object to add sheets to
        """
        metrics = self._extract_metrics_from_report(data)
        metadata = data.get('metadata', {})

        # Create trend analysis sheet
        self._create_trend_analysis_sheet(workbook, metrics, metadata)

    def build_comparison(
        self,
        data: Dict[str, Any],
        workbook: openpyxl.Workbook,
        historical_data: Optional[Dict[str, Any]] = None
    ) -> None:
        """Create comparison template with current vs previous period side-by-side

        Args:
            data: Current period report data from ReportGenerator.generate_report()
            workbook: openpyxl Workbook object to add sheets to
            historical_data: Previous period report data (optional, will show "N/A" if missing)
        """
        current_metrics = self._extract_metrics_from_report(data)
        current_metadata = data.get('metadata', {})

        if historical_data:
            historical_metrics = self._extract_metrics_from_report(historical_data)
            historical_metadata = historical_data.get('metadata', {})
        else:
            historical_metrics = {}
            historical_metadata = {}

        self._create_comparison_sheet(
            workbook,
            current_metrics,
            current_metadata,
            historical_metrics,
            historical_metadata
        )

    # =============================================================================
    # SHEET CREATION METHODS
    # =============================================================================

    def _create_kpi_cards_sheet(
        self,
        workbook: openpyxl.Workbook,
        metrics: Dict[str, Any],
        metadata: Dict[str, Any]
    ) -> None:
        """Create KPI cards sheet with large, colored metric displays"""
        ws = workbook.create_sheet("KPI Dashboard", 0)

        # Title
        ws['A1'] = "Analytics Dashboard"
        ws['A1'].font = Font(bold=True, size=20)
        ws.merge_cells("A1:H1")

        # Date range subtitle
        date_range = metadata.get('date_range', {})
        ws['A2'] = f"Period: {date_range.get('start', 'N/A')} to {date_range.get('end', 'N/A')} ({date_range.get('days', 0)} days)"
        ws['A2'].font = Font(size=12, italic=True)
        ws.merge_cells("A2:H2")

        # KPI cards layout (2 columns, 3 rows)
        kpis = self._get_kpi_metrics(metrics)
        row = 4

        for i in range(0, len(kpis), 2):
            # Left KPI
            if i < len(kpis):
                self._create_kpi_card(ws, kpis[i], row, 1)

            # Right KPI
            if i + 1 < len(kpis):
                self._create_kpi_card(ws, kpis[i + 1], row, 5)

            row += 4  # Space between rows

        # Set column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_charts_sheet(
        self,
        workbook: openpyxl.Workbook,
        metrics: Dict[str, Any],
        metadata: Dict[str, Any]
    ) -> None:
        """Create charts sheet with embedded visualizations"""
        ws = workbook.create_sheet("Charts")

        # Title
        ws['A1'] = "Analytics Charts"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells("A1:H1")

        row = 3

        # Chart 1: Top Skills (Bar Chart)
        top_skills = metrics.get('skills', {}).get('top_skills', [])
        if top_skills:
            row = self._add_bar_chart(
                ws, row,
                title="Top Skills by Invocations",
                data=top_skills,
                label_key='skill_name',
                value_key='invocations'
            )

        # Chart 2: Token Usage by Model (Pie Chart)
        by_model = metrics.get('tokens', {}).get('by_model', {})
        if by_model:
            row = self._add_pie_chart(
                ws, row,
                title="Token Distribution by Model",
                data=by_model
            )

        # Chart 3: Sessions by Day of Week (Bar Chart)
        day_of_week = metrics.get('sessions', {}).get('day_of_week', {})
        if day_of_week:
            row = self._add_bar_chart_from_dict(
                ws, row,
                title="Sessions by Day of Week",
                data=day_of_week
            )

        # Set column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_top_lists_sheet(
        self,
        workbook: openpyxl.Workbook,
        metrics: Dict[str, Any],
        metadata: Dict[str, Any]
    ) -> None:
        """Create top lists sheet with rankings"""
        ws = workbook.create_sheet("Top Lists")

        # Title
        ws['A1'] = "Top Performance Lists"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells("A1:H1")

        row = 3

        # Top Skills
        top_skills = metrics.get('skills', {}).get('top_skills', [])
        if top_skills:
            row = self._add_top_list(
                ws, row,
                title="Top Skills",
                headers=["Rank", "Skill", "Invocations", "Success Rate"],
                data=top_skills,
                columns=['skill_name', 'invocations', 'success_rate']
            )

        # Top Projects (by sessions)
        by_project = metrics.get('sessions', {}).get('by_project', {})
        if by_project:
            # Convert dict to list of tuples
            project_list = sorted(by_project.items(), key=lambda x: x[1], reverse=True)[:10]
            row = self._add_top_list_from_dict(
                ws, row,
                title="Top Projects by Sessions",
                headers=["Rank", "Project", "Sessions"],
                data=project_list
            )

        # Recent Lessons
        recent_lessons = metrics.get('lessons', {}).get('recent_lessons', [])
        if recent_lessons:
            row = self._add_lessons_list(ws, row, recent_lessons)

        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_raw_data_sheet(
        self,
        workbook: openpyxl.Workbook,
        category_name: str,
        data: Dict[str, Any]
    ) -> None:
        """Create raw data sheet for a metric category"""
        ws = workbook.create_sheet(category_name)

        # Title
        ws['A1'] = f"{category_name} - Raw Data"
        ws['A1'].font = Font(bold=True, size=14)

        row = 3

        # Flatten and display all data
        for key, value in data.items():
            ws[f'A{row}'] = key
            ws[f'A{row}'].font = Font(bold=True)

            if isinstance(value, (list, dict)):
                # Complex data types - display as formatted JSON-like structure
                row = self._display_complex_data(ws, row, value)
            else:
                # Simple value
                ws[f'B{row}'] = self._format_value(value)
                row += 1

            row += 1  # Extra spacing

        # Enable auto-filter on all data
        if row > 3:
            ws.auto_filter.ref = f"A3:{get_column_letter(ws.max_column)}{row-1}"

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40

        # Freeze header
        ws.freeze_panes = "A4"

    def _create_trend_analysis_sheet(
        self,
        workbook: openpyxl.Workbook,
        metrics: Dict[str, Any],
        metadata: Dict[str, Any]
    ) -> None:
        """Create trend analysis sheet with time series data"""
        ws = workbook.create_sheet("Trend Analysis", 0)

        # Title
        ws['A1'] = "Trend Analysis"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells("A1:H1")

        # Note: This is a simplified version since we don't have daily time series data
        # In a real implementation, you would need daily/hourly data from collectors

        row = 3

        # Key metrics with trend indicators
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Current Value"
        ws[f'C{row}'] = "Daily Average"
        ws[f'D{row}'] = "Trend"

        # Style header
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.font = Font(bold=True, color=self.COLOR_HEADER_TEXT)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

        # Add trend metrics
        trends = [
            ("Total Sessions", metrics.get('sessions', {}).get('total_sessions', 0), "→"),
            ("Total Skills Used", metrics.get('skills', {}).get('total_invocations', 0), "↑"),
            ("Total Tokens", metrics.get('tokens', {}).get('total_tokens', 0), "↑"),
            ("Success Rate", f"{metrics.get('skills', {}).get('success_rate_overall', 0):.1%}", "→"),
            ("Avg Session Duration", f"{metrics.get('sessions', {}).get('avg_duration_minutes', 0):.1f} min", "↓"),
        ]

        for metric_name, value, trend in trends:
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = value
            ws[f'C{row}'] = "N/A"  # Would need daily data
            ws[f'D{row}'] = trend

            # Color code trend
            if trend == "↑":
                ws[f'D{row}'].font = Font(color=self.COLOR_SUCCESS, bold=True, size=14)
            elif trend == "↓":
                ws[f'D{row}'].font = Font(color=self.COLOR_DANGER, bold=True, size=14)
            else:
                ws[f'D{row}'].font = Font(color=self.COLOR_NEUTRAL, bold=True, size=14)

            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10

        # Freeze header
        ws.freeze_panes = "A4"

    def _create_comparison_sheet(
        self,
        workbook: openpyxl.Workbook,
        current_metrics: Dict[str, Any],
        current_metadata: Dict[str, Any],
        historical_metrics: Dict[str, Any],
        historical_metadata: Dict[str, Any]
    ) -> None:
        """Create comparison sheet with current vs previous period"""
        ws = workbook.create_sheet("Comparison", 0)

        # Title
        ws['A1'] = "Period Comparison"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells("A1:F1")

        # Period labels
        current_range = current_metadata.get('date_range', {})
        historical_range = historical_metadata.get('date_range', {})

        row = 3
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = f"Current Period"
        ws[f'C{row}'] = f"Previous Period"
        ws[f'D{row}'] = "Change"
        ws[f'E{row}'] = "% Change"

        # Style header
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws[f'{col}{row}']
            cell.font = Font(bold=True, color=self.COLOR_HEADER_TEXT)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

        # Add period info
        ws[f'A{row}'] = "Date Range"
        ws[f'B{row}'] = f"{current_range.get('start', 'N/A')} to {current_range.get('end', 'N/A')}"
        ws[f'C{row}'] = f"{historical_range.get('start', 'N/A')} to {historical_range.get('end', 'N/A')}" if historical_range else "N/A"
        ws[f'A{row}'].font = Font(italic=True)
        row += 2

        # Comparison metrics
        comparisons = [
            ("Total Sessions",
             current_metrics.get('sessions', {}).get('total_sessions', 0),
             historical_metrics.get('sessions', {}).get('total_sessions', 0)),
            ("Total Skill Invocations",
             current_metrics.get('skills', {}).get('total_invocations', 0),
             historical_metrics.get('skills', {}).get('total_invocations', 0)),
            ("Total Tokens",
             current_metrics.get('tokens', {}).get('total_tokens', 0),
             historical_metrics.get('tokens', {}).get('total_tokens', 0)),
            ("Total Cost (USD)",
             current_metrics.get('tokens', {}).get('total_cost_usd', 0.0),
             historical_metrics.get('tokens', {}).get('total_cost_usd', 0.0)),
            ("Success Rate",
             current_metrics.get('skills', {}).get('success_rate_overall', 0.0),
             historical_metrics.get('skills', {}).get('success_rate_overall', 0.0)),
            ("Avg Session Duration (min)",
             current_metrics.get('sessions', {}).get('avg_duration_minutes', 0.0),
             historical_metrics.get('sessions', {}).get('avg_duration_minutes', 0.0)),
        ]

        for metric_name, current_val, historical_val in comparisons:
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = self._format_value(current_val)
            ws[f'C{row}'] = self._format_value(historical_val) if historical_val else "N/A"

            # Calculate change
            if historical_val and isinstance(current_val, (int, float)) and isinstance(historical_val, (int, float)):
                change = current_val - historical_val
                pct_change = (change / historical_val * 100) if historical_val != 0 else 0

                ws[f'D{row}'] = self._format_value(change)
                ws[f'E{row}'] = f"{pct_change:.1f}%"

                # Conditional formatting
                if pct_change > 0:
                    ws[f'E{row}'].font = Font(color=self.COLOR_SUCCESS, bold=True)
                    ws[f'E{row}'].value = f"↑ {pct_change:.1f}%"
                elif pct_change < 0:
                    ws[f'E{row}'].font = Font(color=self.COLOR_DANGER, bold=True)
                    ws[f'E{row}'].value = f"↓ {pct_change:.1f}%"
                else:
                    ws[f'E{row}'].font = Font(color=self.COLOR_NEUTRAL)
                    ws[f'E{row}'].value = f"→ 0.0%"
            else:
                ws[f'D{row}'] = "N/A"
                ws[f'E{row}'] = "N/A"

            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        # Freeze header
        ws.freeze_panes = "A5"

    # =============================================================================
    # HELPER METHODS - DATA EXTRACTION
    # =============================================================================

    def _extract_metrics_from_report(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Extract metrics from report data structure

        Report data from ReportGenerator has structure:
        {
            "metadata": {...},
            "sections": [{"title": "...", "metrics": {...}}, ...]
        }

        This flattens it into a single metrics dict organized by category.
        """
        metrics = {
            'skills': {},
            'tokens': {},
            'sessions': {},
            'models': {},
            'lessons': {},
            'workflows': {}
        }

        # Extract from sections
        for section in data.get('sections', []):
            section_metrics = section.get('metrics', {})

            # Map section titles to metric categories
            title = section.get('title', '').lower()

            if 'skill' in title:
                metrics['skills'].update(section_metrics)
            elif 'token' in title or 'cost' in title:
                metrics['tokens'].update(section_metrics)
            elif 'session' in title:
                metrics['sessions'].update(section_metrics)
            elif 'model' in title:
                metrics['models'].update(section_metrics)
            elif 'lesson' in title:
                metrics['lessons'].update(section_metrics)
            elif 'workflow' in title:
                metrics['workflows'].update(section_metrics)
            elif 'overview' in title:
                # Overview section contains mixed metrics
                for key, value in section_metrics.items():
                    if 'skill' in key:
                        metrics['skills'][key] = value
                    elif 'token' in key or 'cost' in key:
                        metrics['tokens'][key] = value
                    elif 'session' in key:
                        metrics['sessions'][key] = value

        return metrics

    def _get_kpi_metrics(self, metrics: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Extract key KPI metrics for dashboard cards"""
        return [
            {
                "label": "Total Sessions",
                "value": metrics.get('sessions', {}).get('total_sessions', 0),
                "format": "number",
                "color": self.COLOR_KPI_BG
            },
            {
                "label": "Total Skills Used",
                "value": metrics.get('skills', {}).get('total_invocations', 0),
                "format": "number",
                "color": self.COLOR_KPI_BG
            },
            {
                "label": "Total Tokens",
                "value": metrics.get('tokens', {}).get('total_tokens', 0),
                "format": "number",
                "color": self.COLOR_KPI_BG
            },
            {
                "label": "Total Cost",
                "value": metrics.get('tokens', {}).get('total_cost_usd', 0.0),
                "format": "currency",
                "color": self.COLOR_WARNING
            },
            {
                "label": "Success Rate",
                "value": metrics.get('skills', {}).get('success_rate_overall', 0.0),
                "format": "percentage",
                "color": self.COLOR_SUCCESS
            },
            {
                "label": "Avg Session Duration",
                "value": metrics.get('sessions', {}).get('avg_duration_minutes', 0.0),
                "format": "duration",
                "color": self.COLOR_KPI_BG
            }
        ]

    # =============================================================================
    # HELPER METHODS - UI COMPONENTS
    # =============================================================================

    def _create_kpi_card(
        self,
        ws: Any,
        kpi: Dict[str, Any],
        row: int,
        col: int
    ) -> None:
        """Create a single KPI card (3x3 merged cell block)"""
        # Label (top row)
        label_cell = ws.cell(row=row, column=col)
        label_cell.value = kpi['label']
        label_cell.font = Font(bold=True, size=11)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)

        # Value (main row - larger)
        value_cell = ws.cell(row=row+1, column=col)

        # Format value based on type
        format_type = kpi.get('format', 'number')
        value = kpi['value']

        if format_type == 'number':
            value_cell.value = value
            value_cell.number_format = '#,##0'
        elif format_type == 'currency':
            value_cell.value = value
            value_cell.number_format = '$#,##0.00'
        elif format_type == 'percentage':
            value_cell.value = value
            value_cell.number_format = '0.0%'
        elif format_type == 'duration':
            value_cell.value = f"{value:.1f} min"
        else:
            value_cell.value = value

        value_cell.font = Font(bold=True, size=18, color=self.COLOR_KPI_TEXT)
        value_cell.fill = PatternFill(start_color=kpi['color'], end_color=kpi['color'], fill_type="solid")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+2)

        # Add border
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for r in range(row, row+3):
            for c in range(col, col+3):
                ws.cell(row=r, column=c).border = border

    def _add_bar_chart(
        self,
        ws: Any,
        row: int,
        title: str,
        data: List[Dict[str, Any]],
        label_key: str,
        value_key: str
    ) -> int:
        """Add a bar chart to the worksheet

        Returns:
            int: Next available row
        """
        # Title
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        # Data table
        ws[f'A{row}'] = "Category"
        ws[f'B{row}'] = "Value"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 1

        data_start_row = row

        for item in data[:10]:  # Top 10
            ws[f'A{row}'] = item.get(label_key, 'N/A')
            ws[f'B{row}'] = item.get(value_key, 0)
            row += 1

        data_end_row = row - 1

        # Create chart
        chart = BarChart()
        chart.title = title
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Category"

        data_ref = Reference(ws, min_col=2, min_row=data_start_row-1, max_row=data_end_row)
        cats_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        ws.add_chart(chart, f'D{data_start_row-2}')

        return row + 15  # Leave space for chart

    def _add_bar_chart_from_dict(
        self,
        ws: Any,
        row: int,
        title: str,
        data: Dict[str, int]
    ) -> int:
        """Add a bar chart from a dictionary

        Returns:
            int: Next available row
        """
        # Convert dict to list format
        data_list = [{"label": k, "value": v} for k, v in data.items()]
        return self._add_bar_chart(ws, row, title, data_list, "label", "value")

    def _add_pie_chart(
        self,
        ws: Any,
        row: int,
        title: str,
        data: Dict[str, Any]
    ) -> int:
        """Add a pie chart to the worksheet

        Returns:
            int: Next available row
        """
        # Title
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        # Data table
        ws[f'A{row}'] = "Category"
        ws[f'B{row}'] = "Value"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 1

        data_start_row = row

        for key, value in data.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1

        data_end_row = row - 1

        # Create chart
        chart = PieChart()
        chart.title = title

        data_ref = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
        cats_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)

        chart.add_data(data_ref)
        chart.set_categories(cats_ref)

        ws.add_chart(chart, f'D{data_start_row-2}')

        return row + 15  # Leave space for chart

    def _add_top_list(
        self,
        ws: Any,
        row: int,
        title: str,
        headers: List[str],
        data: List[Dict[str, Any]],
        columns: List[str]
    ) -> int:
        """Add a top list table

        Returns:
            int: Next available row
        """
        # Title
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        # Headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color=self.COLOR_HEADER_TEXT)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1
        data_start_row = row

        # Data rows
        for rank, item in enumerate(data[:10], start=1):
            ws.cell(row=row, column=1).value = rank

            for col_idx, col_name in enumerate(columns, start=2):
                value = item.get(col_name, 'N/A')
                ws.cell(row=row, column=col_idx).value = self._format_value(value)

            # Zebra striping
            if rank % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col_idx).fill = PatternFill(
                        start_color=self.COLOR_STRIPE_2,
                        end_color=self.COLOR_STRIPE_2,
                        fill_type="solid"
                    )

            row += 1

        return row + 2

    def _add_top_list_from_dict(
        self,
        ws: Any,
        row: int,
        title: str,
        headers: List[str],
        data: List[Tuple[str, Any]]
    ) -> int:
        """Add a top list table from a list of (key, value) tuples

        Returns:
            int: Next available row
        """
        # Title
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        # Headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color=self.COLOR_HEADER_TEXT)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

        # Data rows
        for rank, (key, value) in enumerate(data, start=1):
            ws.cell(row=row, column=1).value = rank
            ws.cell(row=row, column=2).value = key
            ws.cell(row=row, column=3).value = value

            # Zebra striping
            if rank % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col_idx).fill = PatternFill(
                        start_color=self.COLOR_STRIPE_2,
                        end_color=self.COLOR_STRIPE_2,
                        fill_type="solid"
                    )

            row += 1

        return row + 2

    def _add_lessons_list(
        self,
        ws: Any,
        row: int,
        lessons: List[Dict[str, Any]]
    ) -> int:
        """Add recent lessons list

        Returns:
            int: Next available row
        """
        # Title
        ws[f'A{row}'] = "Recent Lessons Learned"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        # Headers
        headers = ["#", "Lesson", "Source", "Status"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color=self.COLOR_HEADER_TEXT)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")

        row += 1

        # Data rows
        for idx, lesson in enumerate(lessons[:10], start=1):
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = lesson.get('lesson', 'N/A')
            ws.cell(row=row, column=3).value = lesson.get('source', 'N/A')
            ws.cell(row=row, column=4).value = lesson.get('status', 'N/A')

            # Zebra striping
            if idx % 2 == 0:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col_idx).fill = PatternFill(
                        start_color=self.COLOR_STRIPE_2,
                        end_color=self.COLOR_STRIPE_2,
                        fill_type="solid"
                    )

            row += 1

        return row + 2

    def _display_complex_data(
        self,
        ws: Any,
        row: int,
        data: Any
    ) -> int:
        """Display complex data (lists, dicts) in a readable format

        Returns:
            int: Next row after data
        """
        if isinstance(data, list):
            for idx, item in enumerate(data[:20], start=1):  # Limit to 20 items
                ws[f'B{row}'] = f"[{idx}] {self._format_value(item)}"
                row += 1
        elif isinstance(data, dict):
            for key, value in list(data.items())[:20]:  # Limit to 20 items
                ws[f'B{row}'] = f"{key}: {self._format_value(value)}"
                row += 1
        else:
            ws[f'B{row}'] = str(data)
            row += 1

        return row

    def _format_value(self, value: Any) -> str:
        """Format a value for display"""
        if isinstance(value, float):
            if 0 < value < 1:
                return f"{value:.2%}"  # Likely a percentage
            else:
                return f"{value:,.2f}"
        elif isinstance(value, int):
            return f"{value:,}"
        elif isinstance(value, bool):
            return "Yes" if value else "No"
        elif isinstance(value, (list, dict)):
            return str(value)[:100]  # Truncate long structures
        else:
            return str(value)


def main():
    """Demo usage of ExcelTemplateBuilder"""
    # This would normally use real report data from ReportGenerator
    # For demo, we'll create sample data matching the expected structure

    sample_report = {
        "metadata": {
            "generated_at": "2026-05-01T10:00:00",
            "report_type": "detailed",
            "date_range": {
                "start": "2026-04-01",
                "end": "2026-04-30",
                "days": 30
            }
        },
        "sections": [
            {
                "title": "Overview",
                "metrics": {
                    "total_sessions": 145,
                    "total_skill_invocations": 423,
                    "total_tokens": 1250000,
                    "total_cost_usd": 15.75,
                    "date_range_days": 30
                }
            },
            {
                "title": "Top Skills",
                "metrics": {
                    "top_skills": [
                        {"skill_name": "dream-studio:core", "invocations": 89, "success_rate": 0.95},
                        {"skill_name": "dream-studio:quality", "invocations": 67, "success_rate": 0.92},
                        {"skill_name": "dream-studio:security", "invocations": 45, "success_rate": 0.88},
                    ],
                    "success_rate_overall": 0.93
                }
            },
            {
                "title": "Token Usage",
                "metrics": {
                    "by_model": {
                        "claude-sonnet-4-5": 850000,
                        "claude-haiku-3-5": 350000,
                        "claude-opus-4": 50000
                    },
                    "daily_average": 41667
                }
            },
            {
                "title": "Session Analytics",
                "metrics": {
                    "by_project": {
                        "dream-studio": 78,
                        "dreamysuite": 45,
                        "analytics": 22
                    },
                    "day_of_week": {
                        "Monday": 25,
                        "Tuesday": 30,
                        "Wednesday": 22,
                        "Thursday": 28,
                        "Friday": 20,
                        "Saturday": 10,
                        "Sunday": 10
                    },
                    "avg_duration_minutes": 45.5
                }
            },
            {
                "title": "Lessons Learned",
                "metrics": {
                    "total_lessons": 12,
                    "recent_lessons": [
                        {"lesson": "Always check for open PRs before pushing", "source": "build", "status": "active"},
                        {"lesson": "Use Glob after subagent writes", "source": "debug", "status": "active"},
                    ]
                }
            }
        ]
    }

    # Create builder
    builder = ExcelTemplateBuilder()

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Build each template type in separate workbooks
    print("Building Summary Dashboard template...")
    wb1 = openpyxl.Workbook()
    if "Sheet" in wb1.sheetnames:
        wb1.remove(wb1["Sheet"])
    builder.build_summary_dashboard(sample_report, wb1)
    wb1.save("C:/Users/Dannis Seay/Downloads/demo_summary_dashboard.xlsx")
    print("  Saved: demo_summary_dashboard.xlsx")

    print("Building Raw Data Dump template...")
    wb2 = openpyxl.Workbook()
    if "Sheet" in wb2.sheetnames:
        wb2.remove(wb2["Sheet"])
    builder.build_raw_data_dump(sample_report, wb2)
    wb2.save("C:/Users/Dannis Seay/Downloads/demo_raw_data_dump.xlsx")
    print("  Saved: demo_raw_data_dump.xlsx")

    print("Building Trend Analysis template...")
    wb3 = openpyxl.Workbook()
    if "Sheet" in wb3.sheetnames:
        wb3.remove(wb3["Sheet"])
    builder.build_trend_analysis(sample_report, wb3)
    wb3.save("C:/Users/Dannis Seay/Downloads/demo_trend_analysis.xlsx")
    print("  Saved: demo_trend_analysis.xlsx")

    print("Building Comparison template...")
    wb4 = openpyxl.Workbook()
    if "Sheet" in wb4.sheetnames:
        wb4.remove(wb4["Sheet"])

    # Create historical data (previous period with lower values)
    historical_report = {
        "metadata": {
            "generated_at": "2026-04-01T10:00:00",
            "report_type": "detailed",
            "date_range": {
                "start": "2026-03-01",
                "end": "2026-03-31",
                "days": 31
            }
        },
        "sections": [
            {
                "title": "Overview",
                "metrics": {
                    "total_sessions": 120,
                    "total_skill_invocations": 350,
                    "total_tokens": 1000000,
                    "total_cost_usd": 12.50,
                }
            },
            {
                "title": "Top Skills",
                "metrics": {
                    "success_rate_overall": 0.90
                }
            },
            {
                "title": "Session Analytics",
                "metrics": {
                    "avg_duration_minutes": 50.0
                }
            }
        ]
    }

    builder.build_comparison(sample_report, wb4, historical_report)
    wb4.save("C:/Users/Dannis Seay/Downloads/demo_comparison.xlsx")
    print("  Saved: demo_comparison.xlsx")

    print("\nAll demo templates created successfully!")
    print("Check C:/Users/Dannis Seay/Downloads/ for the files.")


if __name__ == "__main__":
    main()
